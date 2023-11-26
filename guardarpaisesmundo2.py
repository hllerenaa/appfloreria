import os, sys
from django.core.wsgi import get_wsgi_application

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'appfloreria.settings')

application = get_wsgi_application()
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from area_geografica.models import Pais, Provincia, Ciudad, Parroquia
from appfloreria.settings import BASE_DIR
from django.db import transaction
import openpyxl
import json

# with open(os.path.join(BASE_DIR, 'ecuador.json'), 'r') as json_file:
#     provincias = json.load(json_file)
#     pass

wrkbk = openpyxl.load_workbook(os.path.join(BASE_DIR, "INEC_LATAM.xlsx"))
h = wrkbk["Hoja2"]


if __name__ == '__main__':
    try:
        with transaction.atomic():
            for i in range(3, h.max_row + 1):
                ciudad = str(h.cell(row=i, column=1).value).upper()
                lat = str(h.cell(row=i, column=2).value).upper()
                lng = str(h.cell(row=i, column=3).value).upper()
                pais = str(h.cell(row=i, column=4).value).upper()
                provincia = str(h.cell(row=i, column=7).value).upper()
                print(f'{pais}, {provincia}, {ciudad}')
                if ciudad:
                    if not Pais.objects.filter(status=True, nombre__iexact=pais).exists():
                        pais = Pais.objects.create(nombre=pais, codigoidioma="es")
                    else:
                        pais = Pais.objects.get(nombre__iexact=pais)

                    if not Provincia.objects.filter(status=True, nombre__iexact=provincia, pais_id=pais.id).exists():
                        provincia = Provincia.objects.create(nombre=provincia, pais_id=pais.pk)
                    else:
                        provincia = Provincia.objects.get(status=True, nombre__iexact=provincia, pais_id=pais.id)

                    if not Ciudad.objects.filter(status=True, nombre__iexact=ciudad,
                                                 provincia_id=provincia.pk,).exists():
                        ciudad = Ciudad.objects.create(nombre=ciudad, provincia_id=provincia.pk)
                    else:
                        ciudad = Ciudad.objects.get(nombre=ciudad, provincia_id=provincia.pk)

    except Exception as ex:
        print("Error: "+str(ex))