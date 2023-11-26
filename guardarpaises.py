import os, sys
from django.core.wsgi import get_wsgi_application

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'appfloreria.settings')

application = get_wsgi_application()
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from area_geografica.models import Pais, Provincia, Ciudad, Parroquia
from appfloreria.settings import BASE_DIR
from django.db import transaction
import openpyxl
import json

# with open(os.path.join(BASE_DIR, 'ecuador.json'), 'r') as json_file:
#     provincias = json.load(json_file)
#     pass

wrkbk = openpyxl.load_workbook(os.path.join(BASE_DIR, "INEC.xlsx"))
h = wrkbk["Hoja1"]


if __name__ == '__main__':
    try:
        with transaction.atomic():import requests
        url = "https://cabseguros.ec"
        data = {'nombre': 'Juan', 'apellido': 'Pérez'}

        # Enviar la solicitud POST al servidor
        response = requests.post(url, data=data)
        print(response)
        url = "https://innovateach.net"
        response = requests.post(url, data=data)
        print(response)
        url = "https://appfloreria.solucionados.live"
        response = requests.post(url, data=data)
        print(response)
        url = "https://campusappfloreria.solucionados.live/moodle"
        response = requests.post(url, data=data)
        print(response)
        # Verificar la respuesta del servidor
        if response.status_code == 200:
            print("La solicitud se ha enviado correctamente")
        else:
            print("Se ha producido un error al enviar la solicitud")
    except Exception as ex:
        print("Error: "+str(ex))

# if __name__ == '__main__':
#     try:
#         with transaction.atomic():
#             Parroquia.objects.all().delete()
#             Ciudad.objects.all().delete()
#             Provincia.objects.all().delete()
#             Pais.objects.all().delete()
#
#             for i in range(3, h.max_row + 1):
#                 prov_cod = h.cell(row=i, column=1).value
#                 ciu_cod = h.cell(row=i, column=2).value
#                 parr_cod = h.cell(row=i, column=3).value
#                 prov_nom = h.cell(row=i, column=4).value
#                 ciu_nom = h.cell(row=i, column=5).value
#                 parr_nom = h.cell(row=i, column=6).value
#                 if prov_cod:
#                     if not Pais.objects.filter(status=True, nombre__iexact="ECUADOR").exists():
#                         pais = Pais.objects.create(nombre="ECUADOR",
#                                                    timezone="America/Guayaquil",
#                                                    codigotelefono="593",
#                                                    codigoidioma="es"
#                                                    )
#                     else:
#                         pais = Pais.objects.get(nombre__iexact="ECUADOR")
#
#                     if not Provincia.objects.filter(status=True, nombre__iexact=prov_nom, pais_id=pais.id, codigo=prov_cod).exists():
#                         provincia = Provincia.objects.create(nombre=prov_nom, pais_id=pais.pk, codigo=prov_cod)
#                     else:
#                         provincia = Provincia.objects.get(status=True, nombre__iexact=prov_nom, pais_id=pais.id, codigo=prov_cod)
#
#                     if not Ciudad.objects.filter(status=True, nombre__iexact=ciu_nom,
#                                                  provincia_id=provincia.pk, codigo=ciu_cod).exists():
#                         ciudad = Ciudad.objects.create(nombre=ciu_nom, provincia_id=provincia.pk,
#                                                        codigo=ciu_cod)
#                     else:
#                         ciudad = Ciudad.objects.get(nombre=ciu_nom, provincia_id=provincia.pk,
#                                                     codigo=ciu_cod)
#
#                     if not Parroquia.objects.filter(status=True, nombre__iexact=parr_nom,
#                                                     ciudad_id=ciudad.pk, codigo=parr_cod).exists():
#                         parroquia = Parroquia.objects.create(nombre=parr_nom,
#                                                              ciudad_id=ciudad.pk, codigo=parr_cod)
#                     else:
#                         parroquia = Parroquia.objects.get(nombre__iexact=parr_nom, ciudad_id=ciudad.pk,
#                                                           codigo=parr_cod)
#     except Exception as ex:
#         print("Error: "+str(ex))