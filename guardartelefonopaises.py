import os, sys
from django.core.wsgi import get_wsgi_application

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'appfloreria.settings')

application = get_wsgi_application()
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from area_geografica.models import Pais, Provincia, Ciudad, Parroquia
from appfloreria.settings import BASE_DIR
from django.db import transaction
import openpyxl
import json

# with open(os.path.join(BASE_DIR, 'ecuador.json'), 'r') as json_file:
#     provincias = json.load(json_file)
#     pass

wrkbk = openpyxl.load_workbook(os.path.join(BASE_DIR, "PAISTELEFONO.xlsx"))
h = wrkbk["Hoja1"]


if __name__ == '__main__':
    try:
        with transaction.atomic():
            for i in range(3, h.max_row + 1):
                nombre = h.cell(row=i, column=1).value
                cod = h.cell(row=i, column=2).value
                if nombre:
                    if Pais.objects.filter(status=True, nombre__unaccent__icontains=nombre).exists():
                        print(f'{nombre} {cod}')
                        pais_ = Pais.objects.filter(status=True, nombre__unaccent__icontains=nombre).first()
                        pais_.codigotelefono = cod
                        pais_.save()
    except Exception as ex:
        print("Error: "+str(ex))